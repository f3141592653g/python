# -*- coding:utf-8 _*-
""" 
@author:zhouzongquan
@time: 2018/12/17
@email:919839065@qq.com
@function：
"""
import openpyxl


class Case:
    """
    测试用例类，每个测试用例，实际上就是它的一个实例
    """

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None


class DoExcel:

    def __init__(self, file_name, sheet_name):
        # 异常处理
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row  # 获取最大行数

        cases = []  # 列表，存放所有的测试用例
        for r in range(2, max_row + 1):
            case = Case()  # 实例
            case.case_id = self.sheet.cell(row=r, column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value  # sql
            cases.append(case)

        self.workbook.close()
        return cases  # 返回case列表

    def write_result(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()


